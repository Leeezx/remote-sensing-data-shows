"""Validate and spatially assign reclamation-potential source rows."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime, timezone
import gzip
import hashlib
import json
import math
import os
from pathlib import Path
import shutil
import tempfile
from uuid import uuid4

from openpyxl import load_workbook
try:
    from pyproj import CRS
except ImportError:
    CRS = None
from shapely.geometry import mapping, shape
from shapely.ops import unary_union

from backend.shapefile_geojson import iter_shapefile_geojson_features
from scripts.build_township_chunks import point_in_geometry


EXPECTED_COLUMNS = [
    'longitude', 'latitude',
    'EV', 'optimal_irr', 'optimal_npp', 'optimal_soc',
    'EV.1', 'irr', 'npp', 'soc',
]
REAL_DUPLICATE_EV_COLUMNS = [
    'longitude', 'latitude',
    'EV', 'optimal_irr', 'optimal_npp', 'optimal_soc',
    'EV', 'irr', 'npp', 'soc',
]
NODATA = -999.0
SCHEMA_VERSION = 1
UNIT = 'thousand_usd'
POINT_FIELDS = [
    'longitude', 'latitude',
    'current.reclamationValue', 'current.waterConsumption',
    'current.yieldValue', 'current.soilCarbonValue',
    'future.reclamationValue', 'future.waterConsumption',
    'future.yieldValue', 'future.soilCarbonValue',
]
METRICS = [
    {'field': 'reclamationValue', 'label': '复耕价值', 'unit': UNIT},
    {'field': 'waterConsumption', 'label': '用水消耗', 'unit': UNIT},
    {'field': 'yieldValue', 'label': '产量价值', 'unit': UNIT},
    {'field': 'soilCarbonValue', 'label': '土壤碳价值', 'unit': UNIT},
]
SHAPEFILE_SIDECARS = ('.shp', '.shx', '.dbf', '.prj', '.cpg')
MAX_CHINA_OUTLINE_BYTES = 1_000_000


def validate_wgs84_shapefile(path: str | Path) -> None:
    """Require a Shapefile .prj that resolves to WGS 84 longitude/latitude."""
    shapefile_path = Path(path)
    prj_path = shapefile_path.with_suffix('.prj')
    if not prj_path.is_file():
        raise ValueError(
            f'Shapefile CRS must be EPSG:4326 (WGS84); missing .prj: {prj_path}'
        )
    if CRS is None:
        raise ValueError(
            'Shapefile CRS must be EPSG:4326 (WGS84), but pyproj is unavailable '
            f'to validate {prj_path}'
        )
    try:
        source_crs = CRS.from_wkt(prj_path.read_text(encoding='utf-8-sig'))
    except Exception as exc:
        raise ValueError(
            f'Shapefile CRS must be EPSG:4326 (WGS84); could not parse {prj_path}'
        ) from exc
    if not source_crs.equals(CRS.from_epsg(4326), ignore_axis_order=True):
        raise ValueError(
            'Shapefile CRS must be EPSG:4326 (WGS84); '
            f'{shapefile_path} declares {source_crs.to_string()}'
        )


@dataclass(frozen=True)
class ScenarioMetrics:
    reclamation_value: float
    water_consumption: float
    yield_value: float
    soil_carbon_value: float

    def as_tuple(self) -> tuple[float, float, float, float]:
        return (
            self.reclamation_value,
            self.water_consumption,
            self.yield_value,
            self.soil_carbon_value,
        )


@dataclass(frozen=True)
class SourcePoint:
    longitude: float
    latitude: float
    current: ScenarioMetrics
    future: ScenarioMetrics


@dataclass(frozen=True)
class DemoRegion:
    region_id: str
    name: str
    feature: dict


@dataclass(frozen=True)
class RegionAssignment:
    by_region: dict[str, list[SourcePoint]]
    unassigned_indexes: list[int]
    overlapping_indexes: list[int]


def _number(value, *, row_number: int, column: str) -> float:
    if isinstance(value, bool):
        raise ValueError(f'row {row_number}: {column} must be a finite number')
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'row {row_number}: {column} must be a finite number') from exc
    if not math.isfinite(result):
        raise ValueError(f'row {row_number}: {column} must be a finite number')
    return result


def _scenario(values: tuple[float, float, float, float], *, row_number: int, name: str) -> ScenarioMetrics:
    nodata_values = [value == NODATA for value in values]
    if any(nodata_values) and not all(nodata_values):
        raise ValueError(f'row {row_number}: {name} has mixed -999 values')
    return ScenarioMetrics(*values)


def read_workbook_points(path: str | Path) -> list[SourcePoint]:
    """Read a strictly validated reclamation workbook into source points."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ['pixel_values']:
            raise ValueError('row 1: workbook must contain exactly one pixel_values sheet')
        sheet = workbook['pixel_values']
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if list(header) not in (EXPECTED_COLUMNS, REAL_DUPLICATE_EV_COLUMNS):
            raise ValueError('row 1: pixel_values header must exactly match EXPECTED_COLUMNS')

        points: list[SourcePoint] = []
        coordinates: set[tuple[float, float]] = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) != len(EXPECTED_COLUMNS):
                raise ValueError(f'row {row_number}: expected {len(EXPECTED_COLUMNS)} columns')
            values = tuple(
                _number(value, row_number=row_number, column=column)
                for column, value in zip(EXPECTED_COLUMNS, row)
            )
            longitude, latitude = values[:2]
            if not -180 <= longitude <= 180:
                raise ValueError(f'row {row_number}: longitude must be in [-180, 180]')
            if not -90 <= latitude <= 90:
                raise ValueError(f'row {row_number}: latitude must be in [-90, 90]')
            coordinate = (longitude, latitude)
            if coordinate in coordinates:
                raise ValueError(f'row {row_number}: duplicate coordinate pair')
            coordinates.add(coordinate)
            points.append(SourcePoint(
                longitude=longitude,
                latitude=latitude,
                current=_scenario(values[2:6], row_number=row_number, name='current'),
                future=_scenario(values[6:10], row_number=row_number, name='future'),
            ))
        return points
    finally:
        workbook.close()


def normalize_region_features(features) -> list[DemoRegion]:
    """Normalize WRRCD/WRRNM GeoJSON properties for deterministic assignment."""
    regions: list[DemoRegion] = []
    region_ids: set[str] = set()
    region_names: set[str] = set()
    for index, feature in enumerate(features):
        if not isinstance(feature, dict):
            raise ValueError(f'region {index}: feature must be an object')
        properties = feature.get('properties')
        if not isinstance(properties, dict):
            raise ValueError(f'region {index}: properties must be an object')
        raw_region_id = properties.get('WRRCD')
        raw_name = properties.get('WRRNM')
        region_id = '' if raw_region_id is None else str(raw_region_id).strip()
        name = '' if raw_name is None else str(raw_name).strip()
        if not region_id:
            raise ValueError(f'region {index}: WRRCD must be non-empty')
        if not name:
            raise ValueError(f'region {index}: WRRNM must be non-empty')
        if region_id in region_ids:
            raise ValueError(f'region {index}: duplicate WRRCD {region_id}')
        if name in region_names:
            raise ValueError(f'region {index}: duplicate WRRNM {name}')
        geometry = feature.get('geometry')
        if not isinstance(geometry, dict):
            raise ValueError(f'region {index}: geometry must be an object')
        region_ids.add(region_id)
        region_names.add(name)
        normalized_feature = {
            **feature,
            'properties': {**properties, 'id': region_id, 'name': name},
        }
        regions.append(DemoRegion(region_id, name, normalized_feature))
    return regions


def assign_points(points: list[SourcePoint], regions: list[DemoRegion]) -> RegionAssignment:
    """Assign point centers to every matching region and retain audit indexes."""
    by_region = {region.region_id: [] for region in regions}
    unassigned_indexes: list[int] = []
    overlapping_indexes: list[int] = []
    for index, point in enumerate(points):
        matches = [
            region for region in regions
            if point_in_geometry((point.longitude, point.latitude), region.feature['geometry'])
        ]
        for region in matches:
            by_region[region.region_id].append(point)
        if not matches:
            unassigned_indexes.append(index)
        elif len(matches) > 1:
            overlapping_indexes.append(index)
    return RegionAssignment(by_region, unassigned_indexes, overlapping_indexes)


def compact_point(point: SourcePoint) -> list[float]:
    """Encode one validated source point in the stable transport-field order."""
    return [
        round(point.longitude, 6),
        round(point.latitude, 6),
        *(round(value, 6) for value in point.current.as_tuple()),
        *(round(value, 6) for value in point.future.as_tuple()),
    ]


def encode_json(payload: object) -> bytes:
    """Serialize artifacts deterministically and without whitespace overhead."""
    return json.dumps(
        payload,
        ensure_ascii=False,
        separators=(',', ':'),
        sort_keys=True,
    ).encode('utf-8')


def encode_gzip(raw: bytes) -> bytes:
    """Create a reproducible gzip representation of an artifact."""
    return gzip.compress(raw, compresslevel=6, mtime=0)


def read_demo_regions(path: str | Path) -> list[DemoRegion]:
    """Read and normalize the demo-region Shapefile."""
    shapefile_path = Path(path)
    validate_wgs84_shapefile(shapefile_path)
    return normalize_region_features(iter_shapefile_geojson_features(shapefile_path))


def read_county_features(path: str | Path) -> list[dict]:
    """Read county boundaries as GeoJSON features for the China overview."""
    shapefile_path = Path(path)
    validate_wgs84_shapefile(shapefile_path)
    return list(iter_shapefile_geojson_features(shapefile_path))


def build_china_outline(features, tolerance: float = 0.05) -> dict:
    """Dissolve and topology-preservingly simplify county geometry."""
    geometries = [shape(feature['geometry']) for feature in features if feature.get('geometry')]
    if not geometries:
        raise ValueError('county source contains no polygon geometry')
    outline = unary_union(geometries).simplify(tolerance, preserve_topology=True)
    if outline.geom_type not in {'Polygon', 'MultiPolygon'}:
        raise ValueError(f'county dissolve produced unsupported geometry {outline.geom_type}')
    return mapping(outline)


def _region_bounds(region: DemoRegion) -> list[list[float]]:
    min_longitude, min_latitude, max_longitude, max_latitude = shape(
        region.feature['geometry']
    ).bounds
    return [
        [round(min_latitude, 6), round(min_longitude, 6)],
        [round(max_latitude, 6), round(max_longitude, 6)],
    ]


def _region_feature(region: DemoRegion, point_count: int) -> dict:
    return {
        'type': 'Feature',
        'properties': {
            'id': region.region_id,
            'name': region.name,
            'pointCount': point_count,
            'bounds': _region_bounds(region),
        },
        'geometry': region.feature['geometry'],
    }


def _source_files(workbook: Path, regions_shp: Path, counties_shp: Path) -> list[Path]:
    files = [workbook]
    for shapefile in (regions_shp, counties_shp):
        files.extend(
            shapefile.with_suffix(suffix)
            for suffix in SHAPEFILE_SIDECARS
            if shapefile.with_suffix(suffix).is_file()
        )
    return files


def _source_manifest(files: list[Path]) -> list[dict]:
    result = []
    for source in sorted(files, key=lambda item: item.name):
        result.append({
            'name': source.name,
            'sha256': hashlib.sha256(source.read_bytes()).hexdigest(),
        })
    return result


def _built_at(files: list[Path]) -> str:
    newest_mtime = max(path.stat().st_mtime for path in files)
    return datetime.fromtimestamp(newest_mtime, timezone.utc).isoformat(
        timespec='microseconds'
    ).replace('+00:00', 'Z')


def _write_json(path: Path, payload: object, *, gzip_copy: bool = False) -> bytes:
    raw = encode_json(payload)
    path.write_bytes(raw)
    if gzip_copy:
        path.with_suffix(path.suffix + '.gz').write_bytes(encode_gzip(raw))
    return raw


def _validate_staged_artifacts(stage: Path, point_ids: list[str]) -> None:
    outline = stage / 'china_outline.geojson'
    if outline.stat().st_size >= MAX_CHINA_OUTLINE_BYTES:
        raise ValueError('simplified China outline must be smaller than 1,000,000 bytes')
    gzip_pairs = [stage / 'overview.json'] + [stage / 'points' / f'{region_id}.json' for region_id in point_ids]
    for raw_path in gzip_pairs:
        compressed_path = raw_path.with_suffix(raw_path.suffix + '.gz')
        if gzip.decompress(compressed_path.read_bytes()) != raw_path.read_bytes():
            raise ValueError(f'gzip artifact does not match raw JSON: {raw_path.name}')


def _replace_output(stage: Path, output: Path, force: bool) -> None:
    """Publish a fully validated stage directory without exposing partial artifacts."""
    if output.exists() and not force:
        raise FileExistsError(f'output already exists; pass --force to rebuild: {output}')
    backup = None
    try:
        if output.exists():
            backup = output.with_name(f'.{output.name}.backup-{uuid4().hex}')
            os.replace(output, backup)
        os.replace(stage, output)
    except Exception:
        if backup is not None and backup.exists() and not output.exists():
            os.replace(backup, output)
        raise
    else:
        if backup is not None:
            shutil.rmtree(backup)


def build_reclamation_data(
    workbook: str | Path,
    regions_shp: str | Path,
    counties_shp: str | Path,
    output: str | Path,
    force: bool,
) -> dict:
    """Build complete, deterministic offline artifacts for the reclamation map."""
    workbook_path = Path(workbook)
    regions_path = Path(regions_shp)
    counties_path = Path(counties_shp)
    output_path = Path(output)
    for label, source in (
        ('workbook', workbook_path),
        ('regions Shapefile', regions_path),
        ('counties Shapefile', counties_path),
    ):
        if not source.is_file():
            raise FileNotFoundError(f'{label} not found: {source}')
    if output_path.exists() and not force:
        raise FileExistsError(f'output already exists; pass --force to rebuild: {output_path}')

    points = read_workbook_points(workbook_path)
    regions = sorted(read_demo_regions(regions_path), key=lambda region: region.region_id)
    counties = read_county_features(counties_path)
    assignment = assign_points(points, regions)
    if assignment.overlapping_indexes:
        raise ValueError(f'overlapping point assignments: {assignment.overlapping_indexes}')

    source_files = _source_files(workbook_path, regions_path, counties_path)
    region_features = [
        _region_feature(region, len(assignment.by_region[region.region_id]))
        for region in regions
    ]
    regions_geojson = {'type': 'FeatureCollection', 'features': region_features}
    china_outline = build_china_outline(counties)
    overview = {
        'schemaVersion': SCHEMA_VERSION,
        'unit': UNIT,
        'metrics': METRICS,
        'chinaOutline': china_outline,
        'regions': regions_geojson,
    }
    manifest = {
        'schemaVersion': SCHEMA_VERSION,
        'builtAt': _built_at(source_files),
        'sourceFiles': _source_manifest(source_files),
        'unit': UNIT,
        'metrics': METRICS,
        'fields': POINT_FIELDS,
        'inputPointCount': len(points),
        'assignedPointCount': sum(len(points) for points in assignment.by_region.values()),
        'unassignedPointCount': len(assignment.unassigned_indexes),
        'overlappingPointCount': len(assignment.overlapping_indexes),
        'unassignedIndexes': assignment.unassigned_indexes,
        'regions': [feature['properties'] for feature in region_features],
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    stage = Path(tempfile.mkdtemp(prefix=f'.{output_path.name}.stage-', dir=output_path.parent))
    try:
        points_dir = stage / 'points'
        points_dir.mkdir()
        _write_json(stage / 'regions.geojson', regions_geojson)
        _write_json(stage / 'china_outline.geojson', china_outline)
        _write_json(stage / 'overview.json', overview, gzip_copy=True)
        for region in regions:
            region_id = region.region_id
            if Path(region_id).name != region_id or region_id in {'.', '..'}:
                raise ValueError(f'unsafe region ID for output path: {region_id}')
            payload = {
                'schemaVersion': SCHEMA_VERSION,
                'region': {'id': region_id, 'name': region.name},
                'unit': UNIT,
                'fields': POINT_FIELDS,
                'points': [compact_point(point) for point in assignment.by_region[region_id]],
            }
            _write_json(points_dir / f'{region_id}.json', payload, gzip_copy=True)
        _write_json(stage / 'manifest.json', manifest)
        _validate_staged_artifacts(stage, [region.region_id for region in regions])
        _replace_output(stage, output_path, force)
    except Exception:
        if stage.exists():
            shutil.rmtree(stage)
        raise
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--workbook', type=Path, required=True)
    parser.add_argument('--regions-shp', type=Path, required=True)
    parser.add_argument('--counties-shp', type=Path, required=True)
    parser.add_argument('--output', type=Path, required=True)
    parser.add_argument('--force', action='store_true')
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    manifest = build_reclamation_data(
        args.workbook,
        args.regions_shp,
        args.counties_shp,
        args.output,
        args.force,
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
