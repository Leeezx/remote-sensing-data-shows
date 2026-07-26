import gzip
import json
import shutil

from openpyxl import Workbook, load_workbook
import pytest
from shapely.geometry import Polygon

import scripts.build_reclamation_data as builder

from scripts.build_reclamation_data import (
    EXPECTED_COLUMNS,
    ScenarioMetrics,
    SourcePoint,
    assign_points,
    normalize_region_features,
    read_workbook_points,
)


def write_workbook(path, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'pixel_values'
    sheet.append(EXPECTED_COLUMNS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def make_point(longitude, latitude):
    return SourcePoint(
        longitude,
        latitude,
        ScenarioMetrics(1.0, 2.0, 3.0, 4.0),
        ScenarioMetrics(5.0, 6.0, 7.0, 8.0),
    )


def square_feature(min_x, min_y, max_x, max_y):
    return {
        'type': 'Feature',
        'properties': {},
        'geometry': {
            'type': 'Polygon',
            'coordinates': [[
                [min_x, min_y], [max_x, min_y], [max_x, max_y],
                [min_x, max_y], [min_x, min_y],
            ]],
        },
    }


def square_region(region_id, name):
    feature = square_feature(100, 30, 102, 32)
    feature['properties'] = {'id': region_id, 'name': name}
    return builder.DemoRegion(region_id, name, feature)


def test_read_workbook_maps_both_scenarios_and_rejects_mixed_nodata(tmp_path):
    source = tmp_path / 'values.xlsx'
    write_workbook(source, [[105.0, 38.0, 1, 2, 3, 4, -999, -999, -999, -999]])

    points = read_workbook_points(source)

    assert points[0].current.as_tuple() == (1.0, 2.0, 3.0, 4.0)
    assert points[0].future.as_tuple() == (-999.0, -999.0, -999.0, -999.0)

    write_workbook(source, [[105.0, 38.0, 1, -999, 3, 4, 5, 6, 7, 8]])
    with pytest.raises(ValueError, match='row 2.*mixed -999'):
        read_workbook_points(source)


def test_read_workbook_rejects_extra_sheets(tmp_path):
    source = tmp_path / 'values.xlsx'
    write_workbook(source, [])
    workbook = load_workbook(source)
    workbook.create_sheet('notes')
    workbook.save(source)

    with pytest.raises(ValueError, match='exactly one pixel_values sheet'):
        read_workbook_points(source)


def test_read_workbook_accepts_real_duplicate_ev_future_header(tmp_path):
    source = tmp_path / 'values.xlsx'
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'pixel_values'
    sheet.append([
        'longitude', 'latitude', 'EV', 'optimal_irr', 'optimal_npp', 'optimal_soc',
        'EV', 'irr', 'npp', 'soc',
    ])
    sheet.append([105.0, 38.0, 1, 2, 3, 4, 5, 6, 7, 8])
    workbook.save(source)

    points = read_workbook_points(source)

    assert points[0].future.as_tuple() == (5.0, 6.0, 7.0, 8.0)


def test_normalize_region_features_rejects_duplicate_or_missing_identifiers():
    feature = {
        'type': 'Feature',
        'properties': {'WRRCD': 'A', 'WRRNM': '区域A'},
        'geometry': {'type': 'Polygon', 'coordinates': []},
    }

    duplicate_name = {**feature, 'properties': {'WRRCD': 'B', 'WRRNM': '区域A'}}
    with pytest.raises(ValueError, match='duplicate WRRNM 区域A'):
        normalize_region_features([feature, duplicate_name])

    duplicate_id = {**feature, 'properties': {'WRRCD': 'A', 'WRRNM': '区域B'}}
    with pytest.raises(ValueError, match='duplicate WRRCD A'):
        normalize_region_features([feature, duplicate_id])

    null_id = {**feature, 'properties': {'WRRCD': None, 'WRRNM': '区域A'}}
    with pytest.raises(ValueError, match='WRRCD must be non-empty'):
        normalize_region_features([null_id])

    null_name = {**feature, 'properties': {'WRRCD': 'A', 'WRRNM': None}}
    with pytest.raises(ValueError, match='WRRNM must be non-empty'):
        normalize_region_features([null_name])


def test_assign_points_uses_polygon_centers_and_audits_outside_points():
    regions = normalize_region_features([
        {
            'type': 'Feature',
            'properties': {'WRRCD': 'A', 'WRRNM': '区域A'},
            'geometry': {
                'type': 'Polygon',
                'coordinates': [[[100, 30], [102, 30], [102, 32], [100, 32], [100, 30]]],
            },
        },
    ])
    points = [
        make_point(101.0, 31.0),
        make_point(110.0, 40.0),
    ]

    result = assign_points(points, regions)

    assert [point.longitude for point in result.by_region['A']] == [101.0]
    assert result.unassigned_indexes == [1]
    assert result.overlapping_indexes == []


def test_build_china_outline_preserves_topology_and_stays_compact(monkeypatch):
    calls = []

    class SimplifiedGeometry:
        def simplify(self, tolerance, preserve_topology):
            calls.append((tolerance, preserve_topology))
            return Polygon([(70, 15), (140, 15), (140, 55), (70, 55), (70, 15)])

    monkeypatch.setattr(builder, 'unary_union', lambda _geometries: SimplifiedGeometry())

    outline = builder.build_china_outline([square_feature(70, 15, 140, 55)])

    assert outline['type'] in {'Polygon', 'MultiPolygon'}
    assert calls == [(0.05, True)]
    assert len(json.dumps(outline, separators=(',', ':')).encode('utf-8')) < 1_000_000


def test_build_outputs_compact_tuples_manifest_and_deterministic_gzip(monkeypatch, tmp_path):
    workbook_path = tmp_path / 'values.xlsx'
    regions_path = tmp_path / 'regions.shp'
    counties_path = tmp_path / 'counties.shp'
    output_path = tmp_path / 'output'
    for path in (workbook_path, regions_path, counties_path):
        path.touch()
    monkeypatch.setattr(builder, 'read_workbook_points', lambda _path: [make_point(101, 31)])
    monkeypatch.setattr(builder, 'read_demo_regions', lambda _path: [square_region('A', '区域A')])
    monkeypatch.setattr(builder, 'read_county_features', lambda _path: [square_feature(70, 15, 140, 55)])

    result = builder.build_reclamation_data(
        workbook_path,
        regions_path,
        counties_path,
        output_path,
        force=False,
    )

    payload = json.loads((tmp_path / 'output/points/A.json').read_text(encoding='utf-8'))
    assert payload['points'][0] == [101.0, 31.0, 1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0]
    assert result['inputPointCount'] == 1
    assert result['assignedPointCount'] == 1
    raw = (tmp_path / 'output/points/A.json').read_bytes()
    assert gzip.decompress((tmp_path / 'output/points/A.json.gz').read_bytes()) == raw
    first_gzip = (tmp_path / 'output/points/A.json.gz').read_bytes()

    shutil.rmtree(tmp_path / 'output')
    builder.build_reclamation_data(
        workbook_path,
        regions_path,
        counties_path,
        output_path,
        force=False,
    )
    assert (tmp_path / 'output/points/A.json.gz').read_bytes() == first_gzip


def test_build_rejects_overlaps_without_replacing_existing_output(monkeypatch, tmp_path):
    workbook_path = tmp_path / 'values.xlsx'
    regions_path = tmp_path / 'regions.shp'
    counties_path = tmp_path / 'counties.shp'
    output_path = tmp_path / 'output'
    for path in (workbook_path, regions_path, counties_path):
        path.touch()
    output_path.mkdir()
    (output_path / 'previous.json').write_text('keep me', encoding='utf-8')
    monkeypatch.setattr(builder, 'read_workbook_points', lambda _path: [make_point(101, 31)])
    monkeypatch.setattr(
        builder,
        'read_demo_regions',
        lambda _path: [square_region('A', '区域A'), square_region('B', '区域B')],
    )
    monkeypatch.setattr(builder, 'read_county_features', lambda _path: [square_feature(70, 15, 140, 55)])

    with pytest.raises(ValueError, match='overlapping'):
        builder.build_reclamation_data(
            workbook_path,
            regions_path,
            counties_path,
            output_path,
            force=True,
        )

    assert (output_path / 'previous.json').read_text(encoding='utf-8') == 'keep me'
